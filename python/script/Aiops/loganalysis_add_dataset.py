from openpyxl import load_workbook
import re
import json
import logging
# https://blog.csdn.net/yaos829/article/details/103594988
import requests

brand_map = {
    "科来软件": "klrj",
    "华为": "huawei",
    "绿盟": "lm",
    "华三": "h3c",
    "启明星辰": "qmxc",
    "深信服": "sxf",
    "思科": "sk",
    "捷邦": "jb",
    "华依": 'hy',
    "天融信": "trx",
    "中创信测": "zcxc",
    "迈普": "mp",
    "国盾量子": "gdlz",
    "弘积": "hj",
    "奇安信": "qax",
    '迪普': "dp",
    "宝利通": "blt",
}
device_map = {
    "代理服务器": "proxy",
    "探针设备":"netprobe",
    "波分复用设备":"wdm",
    "DDOS设备":"ddos",
    "IPS设备":"ips",
    "VPN设备":"vpn",
    "负载均衡设备":"loadbalancing",
    "路由器":"router",
    "防火墙":"firewall",
    "交换机":"switch",
    "语音网关":"voicegateway",
    "网络专用服务器":"networkserver",
    "DNS服务器":"dns",
    "虚拟机":"vps",
    "网络应用设备":"networkapp",
    "其他":"other",
    "视频会议设备":"vediomeeting",
    "服务器":"server"
}

data_set_name = set()

def read_excel(tbl_path):
    wb2 = load_workbook(tbl_path)
    ws = wb2["网络设备"] # sheet名
    rows, columns = ws.max_row, ws.columns
    for row in range(2, rows+1):
        brand = ws.cell(row, 3).value
        brand = "other" if not brand else brand
        if brand in brand_map.keys():
            brand = brand_map.get(brand)
        if re.search(r"[\u4e00-\u9fa5]", brand):
            brand = "other"
        brand = brand.lower()

        device_type = ws.cell(row, 4).value
        device_type = "other" if not device_type else device_type

        if device_type in device_map.keys():
            device_type = device_map.get(device_type)
        device_type = device_type.lower()
        if re.search(r"[\u4e00-\u9fa5]", device_type):
            device_type = "other"
        
        data_set_name.add(f"net_syslog_{brand}_{device_type}")

def create_data(data_set_name):
    try:
        url = "http://10.251.52.102:18080/logAnalysis/api/itoa/v1/dataset-manage"

        payload = {
            "databaseType": "esstore",
            "datasetAlias": "网络_"+"_".join(data_set_name.split("_")[1:]),
            "datasetName": data_set_name,
            "sourceType": "custom",
            "messageField": "",
            "remark": "",
            "timestampField": "@collectiontime",
            "localDatasetId": None,
            "centers": [
                {
                    "centerId": 754275092074496,
                    "storageId": 754274834120704
                }
            ],
            "appIds": [0],
            "contextInfo": {
                "filterCondition": "@path,@hostname,@filehashkey",
                "rownumberField": "@rownumber"
            },
            "lifeId": "",
            "userCaredFields": [],
            "category": "datasetManage",
            "datasetId": None
        }
        headers = {
            "Cookie": "UA=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJkdXJhdGlvbiI6MzYwMDAwMCwibGFzdExvZ2luIjoxNzEzODYyMzQwMzU4LCJuYW1lcyI6IltcImpheFwiLFwibG9nQW5hbHlzaXNcIixcImxvZ1NwZWVkXCIsXCJyZWZpbmVyXCJdIiwic2luZ2xlU2lnbk9uIjpmYWxzZSwid2l0aFNlcnZpY2VBdXRoIjoie1wiamF4XCI6dHJ1ZSxcImdhdWdlXCI6dHJ1ZSxcInZpc2lvblwiOnRydWUsXCJkYXRhTW9kZXJuaXphdGlvblwiOnRydWUsXCJkZWFsQW5hbHlzaXNcIjp0cnVlLFwiY21kYlwiOnRydWUsXCJsb2dBbmFseXNpc1wiOnRydWUsXCJsb2dTcGVlZFwiOnRydWUsXCJyZWZpbmVyXCI6dHJ1ZSxcIkFJT3BzXCI6dHJ1ZX0iLCJzZXNzaW9uSWQiOjU0MDEwMDMzMDY0NTE5NjgsInVzZXJOYW1lIjoiYWRtaW4iLCJ1c2VySWQiOiJjWit4ekdpUVBSZWpISW1OZUlKK1FGV2t3Uk9PcTFEelVZZ0FuWmtycTN4dDhzZkMrOGgzK3hLbEpDMmdPb3VqYlZCdTNna29mVUFzYk1aZGpxdEFMNTRzaG5VdThMdkRGa2VSemhJb3VSZWowOGRvSTRsVE5Ud3FoQnF2c2dIeVFmNFY0VzM5UmMzMHkxeUxKVVVBNlRDbTNkN2JuMEw0MVpxQjhTWEJVNE09IiwicHJvZHVjdHMiOiJ7fSJ9.3kOuCyzxcnWFIeD1ts8IOOd5U0Ls7LhSbbcGnCsfxQrNvXk_IPaSB9ZXc4trI_AzQsKvOF4iPjH6sNfxEue4eA",
            "content-type": "application/json"
        }

        res = requests.request("POST", url, json=payload, headers=headers).json()
        if res.get("success"):
            logging.info(f'{data_set_name}创建成功:{res.get("entity").get("datasetId")}"')
        else:
            logging.error(f"{data_set_name}创建失败")
            logging.info(res)
    except Exception as e:
        logging.exception(e)
        logging.error(f"{data_set_name}创建失败")


if __name__ == '__main__':
    LOG_FORMAT = "%(filename)s-:%(message)s"
    fh = logging.FileHandler("./python/script/Aiops/loganalysis_add_dataset.log", encoding='utf-8')
    fh.setLevel(logging.INFO) #level和format可以在basicCofig中设置
    logging.basicConfig(handlers=[fh], level=logging.INFO, format=LOG_FORMAT) 

    # file_name = "./python/script/Aiops/网络设备表20240423.xlsx"
    # read_excel(file_name)

    with open("./python/script/Aiops/data_set_name1.json", "r+") as f:
        data_set_list = json.load(f)
    for data_set in data_set_list:
        create_data(data_set)
