import json
import os
import requests
import logging
import logging.handlers
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font  
from datetime import datetime
import re
from openpyxl import load_workbook, Workbook



def init_logger(name, count=10):
    logger = logging.getLogger()
    log_level = logging.DEBUG
    formatter = logging.Formatter("[%(asctime)s] %(levelname)s (%(filename)s:%(lineno)d) %(message)s")
    log_path = os.path.dirname(os.path.abspath(__file__))
    if not os.path.exists(log_path):
        os.mkdir(log_path)
    log_name = log_path + os.sep + 'logs' + os.sep + name + '.log'
    print("日志保存在%s" % log_name)
    ###
    ch_handler = logging.StreamHandler()
    ch_handler.setLevel(log_level)
    ch_handler.setFormatter(formatter)
    time_handler = logging.handlers.TimedRotatingFileHandler(log_name, when="midnight", interval=1,
                                                             encoding='utf-8', backupCount=count)
    time_handler.setLevel(log_level)
    time_handler.setFormatter(formatter)
    ###
    logger.setLevel(log_level)
    logger.addHandler(time_handler)
    logger.addHandler(ch_handler)
    return logger

all_index_type_set = set() # 指标的类型 {'top'} 
alert_rule1_all_keys_set = set() # 告警规则所有key集合 {'frequency', 'values', 'kpi', 'kpi_min_thr', 'kpis', 'type', 'duration', 'op', 'value'}
alert_rule2_all_keys_set = set() # 告警规则所有key集合 # {'type', 'special_time_rules', 'strategy'}
alert_rule1_type_set = set() # 告警规则类型的集合 {None, 'alert_level'}
alert_rule2_type_set = set() # 告警规则类型的集合  {'lose_data', 'level_frequency', None, 'anomaly_value', 'single_value', 'anomaly_frequency_ignore_lose_data', 'anomaly_frequency'}
alert_rule_strategy_set = set() # 告警规则策略的集合 {'minor', None, 'default', 'major'}


alert_type = {
    "lose_data": "无数据推送告警",  # 无数据推送告警 kpi.aiops_APPID_PMB.GDLF_busTransCount 在 600 秒内无数据推送则触发告警
    'level_frequency': "异常程度告警",  # kpi.aiops_APPID_UBI_transAvgProcTime 在 600秒内出现 9个异常点的异常程度大于等于 3级至少 1个个kpi满足则触发告警"
    'anomaly_value': "异常值告警", # 异常值告警CBCS_busTransCount 在 180秒内出现 3个异常点的值 <2
    'single_value': "固定阈值告警", # 固定阈值告警CBCS_tecSuccRate 的值 <0.95
    'anomaly_frequency_ignore_lose_data': "连续异常点告警", # 连续异常点告警kpi.aiops_APPID|SAPPID|CGF_IPS|IPS.DCPS|1_busTransCount+ 3 连续出现 6个异常点(忽略断点),至少 1个个kpi满足则触发告警 
    'anomaly_frequency': "异常次数告警"  # 异常次数告警CBCS_transAvgProcTime 在 600秒内出现 10个异常点,至少 1个个kpi异常触发告警
    }
alert_level = {
    "default": "默认告警类型",
    "minor": "次要告警",
    "major": "主要告警"
}
alert_op = {
    "lt": "<",
    "gt": ">",
    "lte": "<=",
    "gte": ">="
}


def gen_rule_str(rule_list):
    if not rule_list[0]:
        logger.error("没有配置告警规则")
        return
    #         gen_rule_str(data["alert_rules"])
    def join(rule):
        rule_str = ""
        kpi_str = rule.get('kpi') if rule.get('kpi') else ', '.join(rule.get('kpis'))
        # {','.join(rule['kpis'])}
        if rule.get("type") == "lose_data":
            rule_str = f"{alert_type[rule['type']]}, {kpi_str} 在 {rule['duration']} 内无数据推送则触发告警"
        elif rule.get("type") == "level_frequency":
            rule_str = f"{alert_type[rule['type']]}, {kpi_str}在 {rule['duration']} 秒内出现 {rule['frequency']} 个异常点的异常程度大于等于 {min(rule['values'])} 级至少 {rule['kpi_min_thr']}个kpi满足则触发告警"
        elif rule.get("type") == "anomaly_value":
            rule_str = f"{alert_type[rule['type']]}, {kpi_str}在 {rule['duration']}秒内出现 {rule['frequency']}个异常点的值 {alert_op[rule['op']]} {rule['value']}"
        elif rule.get("type") == "single_value":
            rule_str = f"{alert_type[rule['type']]}, {kpi_str} 的值 {alert_op[rule['op']]} {rule['value']}"
        elif rule.get("type") == "anomaly_frequency":
            rule_str = f"{alert_type[rule['type']]}, {kpi_str} 在 {rule['duration']}秒内出现 {rule['frequency']}个异常点,至少 {rule['kpi_min_thr']}个kpi异常触发告警"
        elif rule.get("type") == "anomaly_frequency_ignore_lose_data":
            rule_str = f"{alert_type[rule['type']]}, {kpi_str} 连续出现 {rule['frequency']}个异常点(忽略断点),至少 {rule['kpi_min_thr']}个kpi满足则触发告警"
            # 异常次数告警CBCS_transAvgProcTime 
        return rule_str
    for rule in rule_list:
        rule_str = join(rule)
        if rule.get("add_alert_rule"):
            rule_str += f"且 {join(rule.get('add_alert_rule'))}"
        logger.error(rule_str)
    return rule_str

def get_child_rule(service_id, type):

    url = f"http://10.251.4.18:8083/api/service/{service_id}/childrens"

    payload = {
        "name": "",
        "type": type
    }
    headers = {
        # "Authorization": "Bearer eyJhbGciOiJIUzUxMiJ9.eyJmdWxsX25hbWUiOiJzdXBlcmFkbWluIiwiZXhwIjoxNzE2ODkwMjg2LCJ1c2VybmFtZSI6ImFpb3BzIiwicm9sZXMiOltdfQ.kxiFjFMgOSYDyjHNfrOm8mjPWP-qUQ1KkLmfNEK6a6GCbRYETM728LRGNrSVMXDXZWBtdRW10dNrfyrCJAc50g",
        # "Username": "ywpt",
        "Authorization": "RYZ2TmkPPfiBcmWI713hWhiE4AQfAdcF",
        "content-type": "application/json"
    }

    response = requests.request("POST", url, json=payload, headers=headers)
    res = response.json()
    if res.get("data"):
        formate_data(res["data"])
    else:
        logger.error(res.get("msg"))
    # for item in res2:
    #     gen_rule_str(item["alert_rules"])
# 15-17
def formate_data(all_data):
    global current_dimension
    all_res = []
    for item in all_data:
        # if item["service_id"] == '5f179561107d7c3d15b8a384':
        #     logger.error(item)
        # if item["service_id"] == '5fc84b39107d7c9c20fa641b':
        #     logger.error(item) 
        # all_index_type_set.add(item.get("type"))
        data = {}
        rule_data = []
        data["name"] = item["name"]
        if item["type"] == "perf":
            current_dimension = item["name"]
        # elif 
        rule_data.append(current_dimension)
        rule_data.append(item["name"])
        data["floor"] = item.get("type")
        data["service_id"] = item["service_id"]
        data["alert_rules"] = []
        alert_rules = [[{}, {}]] if not item.get("alert_rules") else item["alert_rules"]
        for rule in alert_rules:
            # 检查map是否有缺少
            # alert_rule1_all_keys_set.update([x for x in rule[0].keys()])
            # alert_rule2_all_keys_set.update([x for x in rule[1].keys()])
            # alert_rule1_type_set.add(rule[0].get("type"))
            # alert_rule2_type_set.add(rule[1].get("type"))
            # alert_rule_strategy_set.add(rule[0].get("strategy"))
            res = rule[1]
            if rule[0].get("strategy"):
                res["strategy"] = rule[0].get("strategy")
                rule_data.append(alert_level.get(res["strategy"]))
            if len(rule) > 2:
                res["add_alert_rule"] = rule[2]
                res["add_alert_rule"]["strategy"] = rule[0].get("strategy")
            data["alert_rules"].append(res)
            logger.error(f"{data['name']}\n")
            rule_data.append(gen_rule_str([res]))

            if rule[0].get("strategy") != rule[1].get("strategy"):
                raise Exception("please check")

        rules_list.append(rule_data)
        all_res.append(data)
        # 写在前面，一次次解析
        # logger.error(f"{data['name']}\n")
        # gen_rule_str(data["alert_rules"])
        if data["floor"] == "top":
            # current_dimension = ""
            res = get_child_rule(data["service_id"], "perf")
        if data["floor"] == "perf":
            # current_dimension = data["name"]
            res = get_child_rule(data["service_id"], "third_level")
        # else:
        #     logger.error("ok")
    return all_res


def gen_excel(rule_list, file_name):
    wb = Workbook()
    work_sheet = wb.active
    for rule in rule_list:
        if len(rule) > 3:
            for i in range(1, len(rule), 2):
                tmp = []
                tmp.append(rule[0])
                tmp.extend(rule[i:i+2])
                work_sheet.append(tmp)
        else:
            work_sheet.append(rule)
    wb.save(file_name)


def gen_dimension(rule):
    # a = r"((2(5[0-5]|[0-4]\d))|[0-1]?\d{1,2})(\.((2(5[0-5]|[0-4]\d))|[0-1]?\d{1,2})){3}"
    if re.search("APPID|TTID", rule):
        return "交易类型TTID"
    elif re.search("APPID|IP", rule):
        return "主机维度"
    return "应用"
        
def filter_paso(conf):
    res = []
    for item in conf:
        if any([paso for paso in need_paso if paso in item.get("name")]):
            res.append(item)
            continue
        conf.remove(item)
    return res

if __name__ == "__main__":
    logger = init_logger(os.path.basename(__file__).strip('.py'))
    # conf = gen_conf(topic)
    # logger.info(conf)
    # create_analysis(conf)
    # create_storage(conf)


    # logger.error(os.getcwd())
    with open(os.path.join(os.path.dirname(__file__), "biszer_analy.json"), "r", encoding='utf-8') as f:
    # all_data = json.load("./Aiops/biszer.json", data.decode('utf-8')) 
        all_data = json.load(f)
    res_file = os.path.join(os.path.dirname(__file__), "biszer_rule.xlsx")
    # need_paso = ['CBC', 'CAM', 'FSAIM', 'BSMP', 'PDMP', 'CBEC', 'FSTA', 'FSDSR', 'CBDSR', 'ACP', 'BEB', 'DSR', 'EPS', 'RLCS', 'FAM', 'NCFD', 'ESSC', 'IACP', 'FACP', 'CIFW', 'PMT', 'WMBP', 'RLBP', 'CBIB', 'CBCS', 'CBS', 'ECIF', 'BCBI', 'CSI']
    # all_data = filter_paso(all_data)

    # 多
    wb = Workbook()
    # wb = load_workbook(res_file)
    for app in all_data:
        global current_dimension
        current_dimension = "应用级"

        rules_list = []
        formate_data([app])
        work_sheet = wb.create_sheet(app["name"])
        work_sheet.append(["维度", "告警对象", "告警级别", "告警条件"])
        for i in range(1,5):
            work_sheet.cell(row=1,column=i).font = Font(name="Arial", size=14, color="FF000000", bold=True)  
        for rule in rules_list:
            if len(rule) < 4:
                # work_sheet.append(["应用级", rule[0]])
                continue
            if len(rule) > 4:
                for i in range(2, len(rule), 2):
                    tmp = []
                    tmp.extend(rule[0:2])
                    tmp.extend(rule[i:i+2])
                    # tmp.insert(0, gen_dimension(tmp[2]))
                    work_sheet.append(tmp)
            else:
                # rule.insert(0, gen_dimension(rule[2]))
                work_sheet.append(rule)
    ws = wb.active
    wb.remove(ws)
    wb.save(res_file)
    # 单
    # formate_data(all_data)  
    # gen_excel(rules_list, res_file)
    logger.error("ok")
    print("ok")
