"""
读写 .xlsx → openpyxl 或 pandas
数据分析 → pandas
旧版 .xls → xlrd（或转成 .xlsx）
生成复杂报表 → xlsxwriter  （支持.xlsx	）


"""

from openpyxl import load_workbook
# https://blog.csdn.net/yaos829/article/details/103594988

import pandas as pd

def read_excel(tbl_path):
    wb2 = load_workbook(tbl_path)
    ws = wb2["网络设备"] # sheet名
    rows, columns = ws.max_row, ws.columns
    for row in range(2, rows+1):
        brand = ws.cell(row, 3).value
        device_type = ws.cell(row, 4).value
    # ws.max_row # 从1开始
    # ws.max_column
    # ws.cell(row=row, column=column).value
    pass

def pandas_pro():
    # 读取 Excel 文件
    df = pd.read_excel("example.xlsx", sheet_name="Sheet1")  # 返回 DataFrame
    # 查看数据
    print(df.head())  # 显示前 5 行
    print(df["Name"])  # 获取 "Name" 列

    # 写入
    # 创建 DataFrame
    data = {
        "Name": ["Alice", "Bob"],
        "Age": [25, 30]
    }
    df = pd.DataFrame(data)
    # 写入 Excel
    df.to_excel("output.xlsx", index=False)  # `index=False` 不写入行索引

    # 修改
    # 读取 Excel
    df = pd.read_excel("example.xlsx")
    # 修改数据
    df["Age"] = df["Age"] + 1  # 年龄 +1
    # 保存修改
    df.to_excel("modified.xlsx", index=False)

def xlxs2dict(file_path):
    """读取表格，返回字典"""
    import pandas as pd

    df = pd.read_excel(file_path)
    res = df.to_dict(orient='records')
    # orient参数控制DataFrame转换为dict的格式，可选值包括:
    # - 'dict': 列名作为key，每列数据作为value的列表
    # - 'list': 返回列名和数据的嵌套列表
    # - 'series': 每列作为一个Series对象
    # - 'split': 返回{'index':[...], 'columns':[...], 'data':[...]}格式
    # - 'records': 返回[{col1:val1, col2:val2,...}, {...}]格式的记录列表
    # - 'index': 行索引作为key，每行数据作为value的字典


if __name__ == '__main__':
    file_name = "./python/script/Aiops/网络设备表20240423.xlsx"
    read_excel(file_name)
