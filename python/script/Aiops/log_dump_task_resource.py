#encoding:utf-8
import math
import os, sys, time,functools
import json
import requests
import re

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font  
from log_util import logger,get_cookies



def try_except(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            res = func(*args, **kwargs)
            return res
        except Exception as e:
            logger.exception(e)
            logger.error(f"函数{func}执行出错")
    return wrapper
        
def do_request(method, url, headers, payload=None):
    try:
        # headers.update({"Cookie":cookie})
        res = requests.request(method, url=url, headers=headers, json=payload,cookies=cookie)
        res = res.json()
        return res
    except Exception as e:
        logger.exception(e)
        return False

def get_collection_tasks(size=500, name=None):
    headers = {
        "Content-Type": "application/json"
    }
    if name:
        url = f"{host}/logAnalysis/api/itoa/dataset/senior?dataSetName={name}&page=0&from=0&size={size}"
    else:
        url = f"{host}/logAnalysis/api/itoa/dataset/senior?page=0&from=0&size={size}"
    res = do_request("GET", url, headers)
    if res.get("retCode") == "0000":
        logger.info(f"获取所有采集任务成功：{res.get('retMsg')}")
        return res.get("entity")
    else:
        logger.error(f"获取所有采集任务失败：{res.get('retMsg')}")

def get_topology(set_id, type):
    """
    获取该任务的拓扑信息，并收集相关任务的资源信息
    set_id:任务的datasetId
    type: 任务类型 collecting 采集 storing 存储
    collecting
    """
    headers = {
        "Content-Type": "application/json"
    }
    url = f"{host}/logAnalysis/api/itoa/dataset/topology?datasetId={set_id}&type={type}"
   
    res = do_request("GET", url, headers)
    if res.get("retCode") == "0000":
        # logger.info(f"获取任务拓扑成功：{res.get('retMsg')}")
        topology_info[set_id] = res.get("entity")

        res =  res.get("entity")
        c_name,c_res_conf,a_name,a_zj_conf,a_sq_conf,s_name,s_zj_conf,s_sq_conf,s_index_name = "", "", "", "", "", "", "", "", ""
        # dev
        a_conf, s_conf = "", ""
        # 获取采集任务信息
        if res.get("collectings"):
            c_name = res.get("collectings")[0].get("dataSetName")
            c_res_conf = f"{res.get('collectings')[0].get('numPartitions')}"

        # 获取解析任务信息
        if res.get("parsings"):
            a_name = res.get('parsings')[0].get('dataSetName')
            for conf in res.get('parsings')[0].get('centerList'):
                # TODO int str？
                if conf.get("centerId") == 754275092074496: # zj
                    a_zj_conf = f"{conf.get('config').get('instanceNum')}*{conf.get('config').get('taskManagerMemory')}"
                elif conf.get("centerId") == 764284838682624: # sq
                    a_sq_conf =  f"{conf.get('config').get('instanceNum')}*{conf.get('config').get('taskManagerMemory')}"
                elif conf.get("centerId") == 183106926206976: # dev
                    a_conf =  f"{conf.get('config').get('instanceNum')}*{conf.get('config').get('taskManagerMemory')}"   
        # 获取存储任务信息
        # prod
        zj_last_day_count, zj_last_day_doc_size, zj_last_day_count_sec = 0,0,0
        sq_last_day_count, sq_last_day_doc_size, sq_last_day_count_sec = 0,0,0
        zj_last_7day_count, zj_last_7day_doc_size, zj_last_7day_count_sec = 0,0,0
        sq_last_7day_count, sq_last_7day_doc_size, sq_last_7day_count_sec = 0,0,0
        # dev
        last_day_count, last_day_doc_size, last_day_count_sec = 0,0,0
        last_7day_count, last_7day_doc_size, last_7day_count_sec = 0,0,0

        if res.get('storings'):
            s_name = res.get('storings')[0].get('dataSetAlias')
            s_index_name = res.get('storings')[0].get('dataSetName')
            for conf in res.get('storings')[0].get('centerList'):
                if conf.get("centerId") == 754275092074496: # zj
                    s_zj_conf = f"{json.loads(conf.get('config')).get('instanceNum')}*{json.loads(conf.get('config')).get('taskManagerMemory')}"
                    es_info = all_es_info.get(s_index_name,{}).get("zj",{})
                    if es_info.get("last_day"):
                        _, zj_last_day_count, zj_last_day_doc_size, zj_last_day_count_sec=es_info.get("last_day", {}).values()
                        zj_last_7day_count, zj_last_7day_doc_size, zj_last_7day_count_sec = es_info.get("last_7_day", {}).values()
                elif conf.get("centerId") == 764284838682624: # sq
                    s_sq_conf =  f"{json.loads(conf.get('config')).get('instanceNum')}*{json.loads(conf.get('config')).get('taskManagerMemory')}"
                    es_info = all_es_info.get(s_index_name,{}).get("sq",{})
                    if es_info.get("last_day"):
                        _, sq_last_day_count, sq_last_day_doc_size, sq_last_day_count_sec=es_info.get("last_day", {}).values()
                        sq_last_7day_count, sq_last_7day_doc_size, sq_last_7day_count_sec = es_info.get("last_7_day", {}).values()
                elif conf.get("centerId") == 183106926206976: # dev
                    s_conf =  f"{json.loads(conf.get('config')).get('instanceNum')}*{json.loads(conf.get('config')).get('taskManagerMemory')}"
                    es_info = all_es_info.get(s_index_name,{}).get("dev",{})
                    if es_info.get("last_day"):
                        _, last_day_count, last_day_doc_size, last_day_count_sec=es_info.get("last_day", {}).values()
                        last_7day_count, last_7day_doc_size, last_7day_count_sec = es_info.get("last_7_day", {}).values()
        if env == "prod":
            ws.append([c_name,c_res_conf,a_name,a_zj_conf,a_sq_conf,s_name,s_zj_conf,s_sq_conf,s_index_name, zj_last_day_count, zj_last_day_doc_size, zj_last_day_count_sec, zj_last_7day_count, zj_last_7day_doc_size, zj_last_7day_count_sec,sq_last_day_count, sq_last_day_doc_size, sq_last_day_count_sec,sq_last_7day_count, sq_last_7day_doc_size, sq_last_7day_count_sec])
        else:
            ws.append([c_name,c_res_conf,a_name,a_conf,s_name,s_conf,s_index_name, last_day_count, last_day_doc_size, last_day_count_sec, last_7day_count, last_7day_doc_size, last_7day_count_sec])
    else:
        logger.error(f"获取任务{set_id}拓扑失败：{res.get('retMsg')}")

def filter_collection():
    name_pattern = r""
    res = []
    for info in all_collection_list:
        if name_pattern:
            if not re.search(name_pattern, info["dataSetName"]):
                logger.info(f"{info['dataSetName']}不匹配，跳过")
                continue
        res.append(info["dataSetId"])
    return res


# def get_collection_info(task_id, check=False):
    """
    check 如果为True，会输出该任务的一些信息
    """
    try:
        url = f"{host}/logAnalysis/api/itoa/dataset/collectionOutPut/{task_id}"
        headers = {
            "Content-Type": "application/json"
        }
        res = do_request("GET", url, headers)
        if res.get("retCode") == "0000":
            res = res.get("entity")
            if check:
                logger.info(f"{res.get('dataSetAlias')}的"
                            f"内存为{res.get('store')[0].get('taskManagerMemory')},"
                            f"并发数为{res.get('store')[0].get('instanceNum')},"
                            f"es分片数为{res.get('store')[0].get('numPartitions')}")
            return res
    except Exception as e:
        logger.exception(e)
        logger.error(res)


    else:
        logger.error(f"获取采集任务{task_id}失败：{res.get('retMsg')}")

# def update_collection(task_id):
    new_conf = gen_new_conf(task_id)
    url = f"{host}/logAnalysis/api/itoa/dataset/collectionOutPut"
    headers = {
        "Content-Type": "application/json"
    }
    payload = {
        "entity": new_conf,
        "statusReq": {"discardSavepoint": True, "offset": "group"}
    }
    res = do_request("PUT", url, headers,payload)
    if res.get("retCode") == "0000":
        logger.info(f"更新采集任务{new_conf.get('dataSetName')}成功：{res.get('retMsg')}")
    else:
        logger.error(f"更新采集任务{task_id}失败：{res.get('retMsg')}")


# def gen_new_conf(task_id):
#     task_info = get_collection_info(task_id)
#     task_info["filter"]["filtersJson"] = conf_templete["filtersJson"]
#     return task_info
def dump_es_info(file_name="es_index_info.json", es_info=None):
    """
    es_info: dict => {"zj": es_info_dict}
    """
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    if not os.path.isfile(file_path):
        with open(file_path, "w+") as f:
            f.write("{}")
    with open(file_path, "r+") as f:
        all_es_info = json.load(f)
        if es_info:
            f.seek(0)
            all_es_info.update(es_info)
            json.dump(all_es_info, f)
        return all_es_info

if __name__ == "__main__":
    env = "prod"

    conf = {
        "prod":{
            "host": "http://10.251.52.102:18080",
            "user": "admin",
            "passwd": "jQ9HQN6kX4Bl8NiZbVmwbyYth8OxdJhjWN+mGiOOXknQxpEb/qS2ayHGpC0U4WXRE5DOHIICHSoW3uAr5jDFeH4Ls9wro/iwORCu3T2VazLt7S6Q3YBdBJk95lWUydTAfXfCIbnvtG70PYGMkA0lNwdnuwhu1CPACyc3oXvNBwY=",
        }
    }.get(env, {})
    host = conf.get("host")
    cookie = get_cookies(f"{host}/authentication/api/login", conf.get("user"), conf.get("passwd"))

    all_es_info = dump_es_info(file_name="calc_es_res.json")
    # all_collection_list = get_collection_tasks(name="采集-SYS_TDSQL_OVERLAY_APPLOG_SOURCE")
    all_collection_list = get_collection_tasks()
    all_collection_list = filter_collection()

    topology_info = {}
    # wb = load_workbook()
    wb = Workbook()
    
    ws=wb.active
    if env == "prod":
        if "prod" in wb.sheetnames:
            work_sheet = wb["prod"] 
            wb.remove(work_sheet) 
        wb.create_sheet(title="prod")		
        title = ["采集任务名称", "采集任务分区数", "解析任务名", "张江解析任务配置", "石泉解析任务配置", "存储任务名", "张江存储任务配置", "石泉存储任务配置", "索引名","最近一天张江索引条数","最近一天张江索引大小","最近一天张江平均每秒","最近七天张江索引条数","最近七天张江平均每天索引大小","最近七天张江平均每秒","最近一天石泉索引条数","最近一天石泉索引大小","最近一天石泉平均每秒","最近七天石泉平均每天索引条数","最近七天石泉索引大小","最近七天石泉平均每秒"]
    else:
        if "dev" in wb.sheetnames:
            work_sheet = wb["dev"] 
            wb.remove(work_sheet) 
        wb.create_sheet(title="dev")
        title = ["采集任务名称", "采集任务分区数", "解析任务名", "解析任务配置", "存储任务名", "存储任务配置", "索引名","最近一天索引条数","最近一天索引大小","最近一天平均每秒","最近七天索引条数","最近七天平均每天索引大小","最近七天平均每秒"]

    
    ws.append(title)
    for i in range(1, len(title) + 1):
        ws.cell(row=1,column=i).font = Font(name="Arial", size=14, color="FF000000", bold=True)  
    for task_id in all_collection_list:
        # 获取任务信息
        get_topology(task_id, type="collecting")
        # 修改任务
    wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'log_task_info.xlsx'))
