from openpyxl import load_workbook, Workbook
"""
# 使用方法
https://blog.51cto.com/u_16213703/8313878
- 文档https://openpyxl-chinese-docs.readthedocs.io/zh-cn/latest/editing_worksheets.html
"""


def fun3_8_2():
    # 新创建的表
    # wb = Workbook()
    # # ws=wb.active # 创建默认的表为sheet

    # ws = wb.active
    # wb.remove(ws)  # 删除sheet
    # ws=wb.create_sheet("test") # 自定义名字 即使是新创建的表，也会有个默认的sheet，如果在新创建的表执行这个，会有两个sheet，所以还是用上条语句
    # ws.append({'C':'dictc','D':'dictd'})
    # wb.save()
    # 从已有的表
    wb = load_workbook('test3.xlsx')
    wb.create_sheet("test5")
    print(wb.sheetnames)
    work_sheet = wb["test5"] 
    # work_sheet.append([3,4,5,6])
    # work_sheet2.insert_rows(7) # 在第7行前面插入一个空行 行数从1开始
    # work_sheet.insert_rows(3,2)  # 在第3行前面插入两个空行 如果行数小于7行，不插入
    # work_sheet.cell(row=1,column=1).font = Font(name="Arial", size=14, color="00FF0000")   # 设置字体
    work_sheet.insert_rows(2, 1)
    # work_sheet2 = wb.create_sheet(title="test") # 如果有同名的，会把新建的sheet重命名 插入到最后
    # work_sheet2 = wb.create_sheet(title="test", 0) #  插入到最前
    # work_sheet2.append({'A':'dicta','B':'dictb'}) # 会在列名为A的最后一行插入dicta,如果没有A，就会新建一列
    # work_sheet2.append({'C':'dictc','D':'dictd'})
    wb.save("test4.xlsx")



if __name__ == '__main__':
    fun3_8_2()
